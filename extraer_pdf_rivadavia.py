import pdfplumber
import pandas as pd
import re
import os
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

pdf_path = "data/polizas/polizaRIV.pdf"

marcas_compuestas = [
    "MERCEDES BENZ", "ALFA ROMEO", "LAND ROVER", "ROLLS ROYCE", "ASTON MARTIN",
    "CHEVROLET CAPTIVA", "VOLKSWAGEN AMAROK", "GREAT WALL", "MINI COOPER"
]

datos = {
    "Marca": "",
    "Modelo": "",
    "Año": "",
    "Suma Asegurada": "--",
    "Premio": "--",
    "Cláusula de Ajuste": "--",
    "Cobertura": "--",
    "Archivo": os.path.basename(pdf_path)
}

with pdfplumber.open(pdf_path) as pdf:
    texto = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])

    def buscar(patron, multilinea=True):
        flags = re.MULTILINE | re.IGNORECASE if multilinea else re.IGNORECASE
        match = re.search(patron, texto, flags)
        return match.group(1).strip() if match else ""

    # Año
    anio_match = re.search(r"MODELO[:\s]+(\d{4})", texto, re.IGNORECASE)
    if anio_match:
        datos["Año"] = anio_match.group(1)

    # Marca y Modelo
    marca_linea = re.search(r"MARCA[:\s]+(.+?)(?:\n|$)", texto, re.IGNORECASE)
    if marca_linea:
        linea = marca_linea.group(1).strip()
        linea = re.sub(r"ASIENTOS.*", "", linea).strip()
        matched = False
        for marca in marcas_compuestas:
            if linea.upper().startswith(marca):
                datos["Marca"] = marca
                datos["Modelo"] = linea[len(marca):].strip()
                matched = True
                break
        if not matched:
            partes = linea.split()
            datos["Marca"] = partes[0]
            datos["Modelo"] = " ".join(partes[1:]) if len(partes) > 1 else ""

    # Suma Asegurada
    suma_match = re.search(r"Suma máxima por Acontecimiento\s+\$?\s*([0-9.,]+)", texto, re.IGNORECASE)
    if suma_match:
        datos["Suma Asegurada"] = suma_match.group(1)

    # Premio
    premio_match = re.search(r"PREMIO\s+\$?\s*([0-9.,]+)", texto, re.IGNORECASE)
    if premio_match:
        datos["Premio"] = premio_match.group(1)
    else:
        for page in pdf.pages:
            tablas = page.extract_tables()
            for tabla in tablas:
                for fila in tabla:
                    for celda in fila:
                        if celda and re.match(r"\d{5,},\d{2}", celda.strip()):
                            datos["Premio"] = celda.strip()
                            break

   
# Cláusula de Ajuste
    ajuste_match = re.search(r"Ajuste Autom[aá]tico\s*[:\-]?\s*(\d{1,3})\s*%", texto, re.IGNORECASE)
    if ajuste_match:
        datos["Cláusula de Ajuste"] = f"{ajuste_match.group(1)}%"


# Cobertura (sólo el contenido, sin repetir el título ni el plan)
cobertura_match = re.search(
    r"Riesgos Cubiertos y Valores Asegurados.*?\n[-]+\n([\s\S]+?)(?=\n\s*(ADVERTENCIA|CA-CO|CO-EX|Frente de P[oó]liza|$))",
    texto,
    re.IGNORECASE
)
if cobertura_match:
    contenido = cobertura_match.group(1).strip()
    contenido = re.sub(r"\n{2,}", "\n", contenido)
    contenido = re.sub(r"\s{2,}", " ", contenido)

    # Remover la línea del ajuste automático si ya la usamos
    contenido = re.sub(r"CA-CC\s*04\.2\s*Ajuste Automático.*", "", contenido, flags=re.IGNORECASE)
    contenido = re.sub(r"Ajuste Autom[aá]tico\s*[:\-]?\s*\d{1,3}\s*%", "", contenido, flags=re.IGNORECASE)

    datos["Cobertura"] = contenido.strip()



# Guardar en Excel
columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura", "Archivo"]
fila_actual = {col: datos[col] for col in columnas}
nombre_archivo = "rivadavia.xlsx"

if os.path.exists(nombre_archivo):
    df_existente = pd.read_excel(nombre_archivo)
    df = pd.concat([df_existente, pd.DataFrame([fila_actual])], ignore_index=True)
else:
    df = pd.DataFrame([fila_actual])

df.to_excel(nombre_archivo, index=False)

# Formato Excel
wb = load_workbook(nombre_archivo)
ws = wb.active
fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for cell in ws[1]:
    cell.fill = fill

for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    for cell in col:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if col[0].value == "Cobertura":
        ws.column_dimensions[col_letter].width = 60
    else:
        ws.column_dimensions[col_letter].width = max_len + 2

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
    ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

wb.save(nombre_archivo)
print(f"✅ Excel generado o actualizado como {nombre_archivo}")
