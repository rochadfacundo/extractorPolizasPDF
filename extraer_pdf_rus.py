import pdfplumber
import pandas as pd
import re
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

pdf_path = "polizaCompleta138627 (1).pdf"

datos = {
    "Marca": "",
    "Modelo": "",
    "Año": "",
    "Suma Asegurada": "",
    "Premio": "--",
    "Cláusula de Ajuste": "--",
    "Cobertura": "--"
}

with pdfplumber.open(pdf_path) as pdf:
    texto_completo = ""
    for page in pdf.pages:
        texto_completo += page.extract_text() + "\n"

    def buscar(patron, multilinea=True):
        flags = re.MULTILINE | re.IGNORECASE if multilinea else re.IGNORECASE
        resultado = re.search(patron, texto_completo, flags)
        return resultado.group(1).strip() if resultado else ""

    # Marca, Modelo, Año
    match_vehiculo = re.search(r"Marca[:.\s]+([A-Z]+)\s+([A-Z0-9\s.]+)\s+Modelo\s+(\d{4})", texto_completo, re.IGNORECASE)
    if match_vehiculo:
        datos["Marca"] = match_vehiculo.group(1).strip()
        datos["Modelo"] = match_vehiculo.group(2).strip()
        datos["Año"] = match_vehiculo.group(3).strip()

    # Suma Asegurada
    datos["Suma Asegurada"] = buscar(r"Suma Asegurada[:.\s]+\$?\s*([0-9.,]+)")

    # Premio
    datos["Premio"] = buscar(r"Premio\s*[:.]*\s*\$?\s*([0-9.,]+)")

    # Cláusula de ajuste
    clausula = buscar(r"Cláusula\s+([A-Z\-]+\s*\d+\.\d+)")
    datos["Cláusula de Ajuste"] = clausula if clausula else "--"

    # Cobertura
    cobertura_match = re.search(
        r"Coberturas especif\.del riesgo\s*:?[\s\n]*((?:.+\n)+?)(?=(?:Descripción del Riesgo|Uso del vehículo|Suma Asegurada))",
        texto_completo,
        re.IGNORECASE
    )
    datos["Cobertura"] = cobertura_match.group(1).strip() if cobertura_match else "--"

# Exportar a Excel
columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura"]
df = pd.DataFrame([{col: datos[col] for col in columnas}])
nombre_archivo = "rio_uruguay.xlsx"
df.to_excel(nombre_archivo, index=False)

# Formateo de Excel
wb = load_workbook(nombre_archivo)
ws = wb.active

# Fondo verde claro en encabezado
fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
for cell in ws[1]:
    cell.fill = fill

# Ajuste de columnas y estilo centrado
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if col[0].value == "Cobertura":
        ws.column_dimensions[col_letter].width = 60
    else:
        ws.column_dimensions[col_letter].width = max_len + 2

# Ajuste de altura de filas
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
    ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

wb.save(nombre_archivo)
print(f"✅ Excel generado como {nombre_archivo}")
