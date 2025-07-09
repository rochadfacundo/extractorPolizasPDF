import pdfplumber
import pandas as pd
import re
import os
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

pdf_path = "data/polizas/polizaMA.pdf"

datos = {
    "Marca": "",
    "Modelo": "",
    "Año": "",
    "Suma Asegurada": "--",
    "Premio": "--",
    "Cláusula de Ajuste": "--",
    "Cobertura": "--",
    "Archivo": os.path.basename(pdf_path)
}

with pdfplumber.open(pdf_path) as pdf:
    texto_completo = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])

    def buscar(patron, multilinea=True):
        flags = re.MULTILINE | re.IGNORECASE if multilinea else re.IGNORECASE
        resultado = re.search(patron, texto_completo, flags)
        return resultado.group(1).strip() if resultado else ""

    # Marca, Modelo, Año (si se puede)
    marca_modelo = buscar(r"Marca.*?: ([^\n]+)")
    if marca_modelo:
        partes = marca_modelo.strip().split(" ")
        datos["Marca"] = partes[0]
        datos["Modelo"] = " ".join(partes[1:-1]) if len(partes) > 2 else partes[1] if len(partes) > 1 else ""
        datos["Año"] = partes[-1] if partes[-1].isdigit() else ""

    # Premio (refacturado o no)
    premio = buscar(r"PREMIO TOTAL\s+([0-9.]+,[0-9]{2})")
    if not premio:
        premio = buscar(r"Cuota\s+Vto\.Asegu\..*?\n\s*1\s+[0-9.]+\s+([0-9.,]+)")
    if not premio:
        premio = buscar(r"Importe\s*\n([0-9.]+,[0-9]{2})")
    if not premio:
        premio = buscar(r"Premio\s*[:]*\s*\$?\s*([0-9.,]+)")
    datos["Premio"] = premio if premio else "--"

# Guardar en Excel (concatenar si existe)
columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura", "Archivo"]
fila_actual = {col: datos[col] for col in columnas}
nombre_archivo = "mercantil.xlsx"

if os.path.exists(nombre_archivo):
    df_existente = pd.read_excel(nombre_archivo)
    df = pd.concat([df_existente, pd.DataFrame([fila_actual])], ignore_index=True)
else:
    df = pd.DataFrame([fila_actual])

df.to_excel(nombre_archivo, index=False)

# Formato de Excel
wb = load_workbook(nombre_archivo)
ws = wb.active
fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for cell in ws[1]:
    cell.fill = fill

for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    for cell in col:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[col_letter].width = 60 if col[0].value == "Cobertura" else max_len + 2

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
    ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

wb.save(nombre_archivo)
print(f"✅ Excel generado o actualizado como {nombre_archivo}")
