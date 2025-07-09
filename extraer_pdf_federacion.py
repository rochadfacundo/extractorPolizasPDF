import pdfplumber
import pandas as pd
import re
import os
import json
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

pdf_path = "data/polizas/polizaFed.pdf"
nombre_excel = "federacion_patronal.xlsx"

datos = {
    "Marca": "--",
    "Modelo": "--",
    "Año": "--",
    "Suma Asegurada": "--",
    "Premio": "--",
    "Cláusula de Ajuste": "--",
    "Cobertura": "--",
    "Archivo": os.path.basename(pdf_path)
}

# Cargar lista de marcas
with open("assets/marcasFiltradasId.json", "r", encoding="utf-8") as f:
    marcas_data = json.load(f)
lista_marcas = [m["marca"].upper() for m in marcas_data]
lista_marcas_ordenadas = sorted(lista_marcas, key=lambda x: len(x.split()), reverse=True)

with pdfplumber.open(pdf_path) as pdf:
    texto_completo = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])

    def buscar(patron, texto=texto_completo):
        match = re.search(patron, texto, re.IGNORECASE)
        return match.group(1).strip() if match else "--"

    # ------- Año -------
    datos["Año"] = buscar(r"Modelo\s+[^\n]*\n.*\b(\d{4})\b")

    # ------- Marca y Modelo -------
    modelo_match = re.search(r"Modelo\s+[^\n]*\n([^\n]+)", texto_completo, re.IGNORECASE)
    if modelo_match:
        linea = modelo_match.group(1).strip().upper()
        for marca in lista_marcas_ordenadas:
            if linea.startswith(marca):
                datos["Marca"] = marca.title()
                datos["Modelo"] = linea[len(marca):].strip().title()
                break

        # Intentar extraer el año desde el final del modelo
        anio_match = re.search(r"(19|20)\d{2}$", datos["Modelo"])
        if anio_match:
            datos["Año"] = anio_match.group(0)
            datos["Modelo"] = re.sub(r"\s+(19|20)\d{2}$", "", datos["Modelo"])
        # ------- Suma Asegurada (tomar la más alta si hay varias) -------
    

    # Buscar Suma Asegurada: buscar la mayor suma con formato 000.000.000,00
    suma_match = re.findall(r"([0-9.]{6,},[0-9]{2})", texto_completo)
    if suma_match:
        # Asumo que la más alta es la Suma Asegurada
        posibles_sumas = [s for s in suma_match if s.count('.') >= 2]
        datos["Suma Asegurada"] = max(posibles_sumas, key=lambda x: float(x.replace('.', '').replace(',', '.'))) if posibles_sumas else "--"
    else:
        datos["Suma Asegurada"] = "--"

    # Buscar Premio: buscar montos con menos cantidad de dígitos
    premio_match = re.findall(r"([0-9.]{1,6},[0-9]{2})", texto_completo)
    if premio_match:
        posibles_premios = [p for p in premio_match if float(p.replace('.', '').replace(',', '.')) < 999999]
        datos["Premio"] = max(posibles_premios, key=lambda x: float(x.replace('.', '').replace(',', '.'))) if posibles_premios else "--"
    else:
        datos["Premio"] = "--"

    # ------- Cláusula de Ajuste -------
    ajuste = buscar(r"Ajuste Autom[aá]tico.*?([0-9]{1,3}\s*%)")
    datos["Cláusula de Ajuste"] = ajuste if ajuste != "--" else "--"

    # ------- Cobertura: línea debajo de "Riesgos Cubiertos" -------
# Buscar línea con "RIESGOS CUBIERTOS" o similares
cobertura_match = re.search(r"RIESGOS CUBIERTOS.*?(\n.*?)(?=\n[A-Z ]+|$)", texto_completo, re.IGNORECASE | re.DOTALL)
if cobertura_match:
    cobertura = cobertura_match.group(1).strip()
    # Limpiar si viene con muchas líneas o info redundante
    cobertura = re.sub(r"\s{2,}", " ", cobertura.replace("\n", " "))
    datos["Cobertura"] = cobertura
else:
    datos["Cobertura"] = "--"
    
    # ---------- EXPORTAR A EXCEL ----------
columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura", "Archivo"]
fila = {col: datos[col] for col in columnas}

if os.path.exists(nombre_excel):
    df_existente = pd.read_excel(nombre_excel)
    df = pd.concat([df_existente, pd.DataFrame([fila])], ignore_index=True)
else:
    df = pd.DataFrame([fila])

df.to_excel(nombre_excel, index=False)

# ---------- ESTÉTICA EXCEL ----------
wb = load_workbook(nombre_excel)
ws = wb.active
fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

for cell in ws[1]:
    cell.fill = fill

for col in ws.columns:
    col_letter = get_column_letter(col[0].column)
    max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    for cell in col:
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.column_dimensions[col_letter].width = 60 if col[0].value == "Cobertura" else max_len + 2

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
    ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

wb.save(nombre_excel)
print(f"✅ Excel generado correctamente como '{nombre_excel}'")
