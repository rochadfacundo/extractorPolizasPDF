import pdfplumber
import pandas as pd
import re
import os
import json
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# --- helpers cobertura Rivadavia ---
def _compact_line(s: str) -> str:
    s = (s or "").replace("\xa0", " ")  # NBSP -> espacio normal
    return re.sub(r"\s+", " ", s.strip())

def _cobertura_rivadavia(texto: str) -> str:
    """
    Devuelve la línea 'Riesgos Cubiertos y Valores Asegurados  <PLAN ...>'
    tomando exactamente el título + lo que sigue en esa misma línea.
    """
    pat = r"(Riesgos\s+Cubiertos\s+y\s+Valores\s+Asegurados)\s+([^\n\r]+)"
    m = re.search(pat, texto, re.IGNORECASE)
    if not m:
        return "--"
    titulo = "Riesgos Cubiertos y Valores Asegurados"
    plan = _compact_line(m.group(2))
    # Limpieza suave si la línea termina en guiones sueltos
    plan = re.sub(r"\s*[-–—_]{2,}\s*$", "", plan)
    return f"{titulo}  {plan}"

# Leer archivo PDF externo
def procesar_rivadavia(pdfs: list[str]):
    # Cargar marcas compuestas desde assets
    with open("assets/marcas.json", encoding="utf-8") as f:
        marcas_data = json.load(f)
    marcas_ordenadas = sorted([m["marca"].upper() for m in marcas_data], key=len, reverse=True)

    filas = []

    for pdf_path in pdfs:
        datos = {
            "Marca": "",
            "Modelo": "",
            "Año": "",
            "Suma Asegurada": "--",
            "Premio": "--",
            "Cláusula de Ajuste": "--",
            "Cobertura": "--",
            "Archivo": os.path.basename(pdf_path)
        }

        with pdfplumber.open(pdf_path) as pdf:
            texto = "\n".join([(p.extract_text() or "") for p in pdf.pages])

            def buscar(patron, multilinea=True):
                flags = re.MULTILINE | re.IGNORECASE if multilinea else re.IGNORECASE
                match = re.search(patron, texto, flags)
                return match.group(1).strip() if match else ""

            # Año
            datos["Año"] = buscar(r"MODELO[:\s]+(\d{4})")

            # Marca y Modelo
            marca_linea = buscar(r"MARCA[:\s]+(.+?)(?:\n|$)")
            if marca_linea:
                linea = re.sub(r"ASIENTOS.*", "", marca_linea).strip().upper()
                for marca in marcas_ordenadas:
                    if linea.startswith(marca):
                        datos["Marca"] = marca.title()
                        datos["Modelo"] = linea[len(marca):].strip().title()
                        break
                else:
                    partes = linea.split()
                    datos["Marca"] = partes[0].title()
                    datos["Modelo"] = " ".join(partes[1:]).title() if len(partes) > 1 else ""

            # Suma Asegurada
            datos["Suma Asegurada"] = buscar(r"Suma máxima por Acontecimiento\s+\$?\s*([0-9.,]+)")

            # Premio
            premio_match = re.search(r"PREMIO\s+\$?\s*([0-9.,]+)", texto, re.IGNORECASE)
            if premio_match:
                datos["Premio"] = premio_match.group(1)
            else:
                # Fallback: buscar en tablas
                for page in pdf.pages:
                    tablas = page.extract_tables()
                    for tabla in tablas or []:
                        for fila in tabla or []:
                            for celda in fila or []:
                                if celda and re.match(r"\d{5,},\d{2}", str(celda).strip()):
                                    datos["Premio"] = str(celda).strip()
                                    break

            # Cláusula de Ajuste
            ajuste = buscar(r"Ajuste Autom[aá]tico\s*[:\-]?\s*(\d{1,3})\s*%")
            if ajuste:
                datos["Cláusula de Ajuste"] = f"{ajuste}%"

            # Cobertura (solo la línea título + plan)
            datos["Cobertura"] = _cobertura_rivadavia(texto)

        filas.append(datos)

    # Guardar en Excel
    columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura", "Archivo"]
    nombre_archivo = "rivadavia.xlsx"

    df_nuevo = pd.DataFrame(filas, columns=columnas)
    if os.path.exists(nombre_archivo):
        df_existente = pd.read_excel(nombre_archivo)
        df = pd.concat([df_existente, df_nuevo], ignore_index=True)
    else:
        df = df_nuevo

    df.to_excel(nombre_archivo, index=False)

    # Formato Excel
    wb = load_workbook(nombre_archivo)
    ws = wb.active
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for cell in ws[1]:
        cell.fill = fill

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = 60 if col[0].value == "Cobertura" else max_len + 2

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
        ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

    wb.save(nombre_archivo)
    print(f"✅ Excel generado o actualizado como {nombre_archivo}")
