import pdfplumber
import pandas as pd
import re
import os
import json
from collections import Counter
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from utils.resources import asset_path,excel_output_path

def _to_float_ar(num_str: str) -> float:
    """Convierte '231.766,69' -> 231766.69 para comparar."""
    return float(num_str.replace('.', '').replace(',', '.'))

def procesar_federacion(pdf_paths: list[str]):
    # === Cargar assets ===
    with open(asset_path("assets", "marcas.json"), "r", encoding="utf-8") as f:
        marcas_data = json.load(f)

    with open(asset_path("assets", "planesFederacion.json"), "r", encoding="utf-8") as f:
        planes_federacion = json.load(f) 

    lista_marcas = [m["marca"].upper() for m in marcas_data]
    lista_marcas_ordenadas = sorted(lista_marcas, key=lambda x: len(x.split()), reverse=True)

    columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura", "Archivo"]
    filas = []

    for pdf_path in pdf_paths:
        datos = dict.fromkeys(columnas, "--")
        datos["Archivo"] = os.path.basename(pdf_path)

        with pdfplumber.open(pdf_path) as pdf:
            texto = "\n".join([(p.extract_text() or "") for p in pdf.pages])

            def buscar(patron):
                m = re.search(patron, texto, re.IGNORECASE)
                return m.group(1).strip() if m else "--"

            # Año (línea de Modelo o al final del modelo)
            datos["Año"] = buscar(r"Modelo\s+[^\n]*\n.*?\b(19|20)\d{2}\b")
            # Marca y Modelo
            modelo_match = re.search(r"Modelo\s+[^\n]*\n([^\n]+)", texto, re.IGNORECASE)
            if modelo_match:
                linea = modelo_match.group(1).strip().upper()
                for marca in lista_marcas_ordenadas:
                    if linea.startswith(marca):
                        datos["Marca"] = marca.title()
                        datos["Modelo"] = linea[len(marca):].strip().title()
                        break
                # Si el año está pegado al final del modelo, separarlo
                anio_match = re.search(r"(19|20)\d{2}$", datos["Modelo"])
                if anio_match:
                    datos["Año"] = anio_match.group(0)
                    datos["Modelo"] = re.sub(r"\s+(19|20)\d{2}$", "", datos["Modelo"]).strip()

            # Suma Asegurada (toma el valor más frecuente)
            suma_matches = re.findall(r"SUMA ASEGURADA\s*\$?\s*([\d.,]+)", texto, re.IGNORECASE)
            if suma_matches:
                datos["Suma Asegurada"] = Counter(suma_matches).most_common(1)[0][0]

            # Premio
            premio_match = re.search(r"PREMIO\s+DEL\s+ENDOSO\s*-?\$?\s*(-?[\d.,]+)", texto, re.IGNORECASE)
            if premio_match:
                datos["Premio"] = premio_match.group(1).lstrip("-")
            else:
                posibles = re.findall(r"(\d{1,6}[.,]\d{2})", texto)
                candidatos = [p for p in posibles if _to_float_ar(p) < 999999]
                if candidatos:
                    datos["Premio"] = max(candidatos, key=_to_float_ar)

            # Cláusula de Ajuste
            datos["Cláusula de Ajuste"] = buscar(r"Ajuste Autom[aá]tico.*?(\d{1,3}\s*%)")

            # Cobertura (plan): normalizar y buscar en el texto
            texto_norm = texto.upper().replace("\n", " ")
            texto_norm = re.sub(r"[\s\-]+", " ", texto_norm)
            plan_encontrado = None
            for plan in planes_federacion:
                plan_norm = re.sub(r"[\s\-]+", " ", plan.upper())
                if plan_norm in texto_norm:
                    plan_encontrado = plan
                    break

            if plan_encontrado:
                datos["Cobertura"] = plan_encontrado
            else:
                # Fallback: bloque de riesgos cubiertos
                m_cov = re.search(r"RIESGOS CUBIERTOS.*?(\n.*?)(?=\n[A-Z ]+|$)", texto, re.IGNORECASE | re.DOTALL)
                if m_cov:
                    datos["Cobertura"] = re.sub(r"\s{2,}", " ", m_cov.group(1).replace("\n", " ")).strip()

        filas.append({col: datos[col] for col in columnas})

    # Excel
    df = pd.DataFrame(filas)
    nombre_excel = excel_output_path("federacion.xlsx")
    df.to_excel(nombre_excel, index=False)

    wb = load_workbook(nombre_excel)
    ws = wb.active
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for cell in ws[1]:
        cell.fill = fill

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = 60 if col[0].value == "Cobertura" else max_len + 2

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
        ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

    wb.save(nombre_excel)
    print(f"✅ Excel generado correctamente como '{nombre_excel}'")
