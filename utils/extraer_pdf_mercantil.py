import pdfplumber
import pandas as pd
import re
import os
import json
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from utils.resources import asset_path,excel_output_path

def _extraer_plan_mercantil(texto: str) -> str:
    """
    En el bloque 'Coberturas especif.del riesgo' toma SOLO la línea de plan,
    p.ej.: 'M PLUS-RCL INC/ROB.TOT.Y PAR/ACC.TOT', 'A - RESPONSABILIDAD CIVIL LIMITADA', etc.
    """
    m = re.search(
        r"Coberturas especif\.del riesgo\s*\n([\s\S]+?)\n\s*Descripci[oó]n del Riesgo",
        texto, re.IGNORECASE
    )
    if not m:
        return ""
    bloque = m.group(1)
    lineas = [ln.strip() for ln in bloque.splitlines() if ln and ln.strip()]

    plan = ""
    despues_de_rc = False
    for ln in lineas:
        if re.search(r"^Responsabilidad\s+civil", ln, re.IGNORECASE):
            despues_de_rc = True
            continue
        if despues_de_rc:
            # Línea candidata: toda en MAYÚSCULAS y que no sea un detalle (Daños/Granizo/etc.)
            if ln.upper() == ln and not re.match(
                r"^(DAÑOS|GRANIZO|ROBO|HUELGA|TERREMOTO|INUNDACI[ÓO]N)", ln, re.IGNORECASE
            ):
                plan = ln
                break

    # Fallbacks
    if not plan:
        cand = re.search(
            r"(?mi)^(?:[ABCMDR]\s*(?:-|–)\s*)?RESPONSABILIDAD CIVIL LIMITADA[^\n]*$",
            bloque
        )
        if cand:
            plan = cand.group(0).strip()

    if not plan:
        cand = re.search(
            r"(?mi)^(?:M\s*(?:PLUS|BAS\.?|BASE|PREMIUM)|B[-0]|B\s*[-.]|C\s*\d?|M\s*BAS\.)[^\n]*RCL[^\n]*$",
            bloque
        )
        if cand:
            plan = cand.group(0).strip()

    return plan

def procesar_mercantil(pdfs: list[str]):
    with open(asset_path("assets", "marcas.json"), "r", encoding="utf-8") as f:
        marcas_data = json.load(f)

    lista_marcas = sorted(
        [m["marca"].upper() for m in marcas_data],
        key=lambda x: len(x.split()),
        reverse=True
    )

    columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura", "Archivo"]
    filas = []

    for pdf_path in pdfs:
        datos = {
            "Marca": "",
            "Modelo": "",
            "Año": "",
            "Suma Asegurada": "--",
            "Premio": "--",
            "Cláusula de Ajuste": "--",
            "Cobertura": "--",
            "Archivo": os.path.basename(pdf_path)
        }

        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = "\n".join([p.extract_text() or "" for p in pdf.pages])

            # buscar() con fuente opcional (por defecto usa texto_completo)
            def buscar(patron, fuente=None, multilinea=True):
                txt = texto_completo if fuente is None else fuente
                flags = re.MULTILINE | re.IGNORECASE if multilinea else re.IGNORECASE
                r = re.search(patron, txt, flags)
                return r.group(1).strip() if r else ""

            # ---------------- Bloque: Descripción del Riesgo ----------------
            bloque_riesgo_m = re.search(
                r"Descripci[oó]n del Riesgo\s*([\s\S]+?)(?=\n\s*(Anexos|Plan de Pago|Prima|Cobert\.|Cl[aá]usulas|$))",
                texto_completo, re.IGNORECASE
            )
            vehiculo_tipo = ""
            marca_linea_riesgo = ""
            anio_riesgo = ""

            def _clean_modelo(s: str) -> str:
                s = re.sub(r"^S/?DESCR\.?\s*", "", s or "", flags=re.IGNORECASE).strip()
                return re.sub(r"\s+", " ", s)

            if bloque_riesgo_m:
                bloque_riesgo = bloque_riesgo_m.group(1)
                vehiculo_tipo = buscar(r"Veh[íi]culo[:.\s]+([^\n]+)", bloque_riesgo)

                # Ejemplo habitual: "Marca: S/DESCR. ...   Modelo: 2009"
                mm = re.search(r"Marca[:.\s]+(.+?)\s+Modelo[:.\s]+(\d{4})", bloque_riesgo, re.IGNORECASE)
                if mm:
                    marca_linea_riesgo = mm.group(1).strip()
                    anio_riesgo = mm.group(2)
                else:
                    # Variantes en líneas separadas
                    marca_linea_riesgo = buscar(r"Marca[:.\s]+([^\n]+)", bloque_riesgo)
                    anio_riesgo = buscar(r"(?:Modelo|A[ÑN]O(?:\s*FABRICACI[ÓO]N)?)[:.\s]+(\d{4})", bloque_riesgo)

            # ---------------- Marca / Modelo / Año (método clásico) ----------------
            marca_modelo = buscar(r"Marca.*?:\s*([^\n]+)")
            if marca_modelo:
                texto = marca_modelo.upper()
                for marca in lista_marcas:
                    if texto.startswith(marca):
                        datos["Marca"] = marca.title()
                        restante = texto[len(marca):].strip()
                        partes = restante.split()
                        if partes and partes[-1].isdigit():
                            datos["Año"] = partes[-1]
                            partes = partes[:-1]
                        datos["Modelo"] = " ".join(partes).title()
                        break

            # ---------------- Reglas especiales: CUATRICICLO / TRAILER ----------------
            if vehiculo_tipo:
                vt = vehiculo_tipo.upper()
                if "CUATRICICL" in vt:
                    datos["Marca"] = "Cuatriciclo"
                    if marca_linea_riesgo:
                        datos["Modelo"] = _clean_modelo(marca_linea_riesgo)  # preserva MAYÚSCULAS
                    if anio_riesgo:
                        datos["Año"] = anio_riesgo
                elif any(k in vt for k in ["TRAILER", "TRÁILER", "REMOLQUE", "ACOPLADO"]):
                    datos["Marca"] = "Trailer"
                    if marca_linea_riesgo:
                        datos["Modelo"] = _clean_modelo(marca_linea_riesgo)
                    if anio_riesgo:
                        datos["Año"] = anio_riesgo

            # ---------------- Premio ----------------
            premio = (
                buscar(r"PREMIO\s*TOTAL\s*\$?\s*([0-9.]+,[0-9]{2})") or
                buscar(r"Prima\s*:?\s*\$?\s*[0-9.,]+\s+Premio\s*:?\s*\$?\s*([0-9.]+,[0-9]{2})") or
                buscar(r"Plan de Pago.*?\n\s*1\s+[0-9./-]+\s*([0-9.]+,[0-9]{2})") or
                buscar(r"Premio\s*[:]*\s*\$?\s*([0-9.]+,[0-9]{2})")
            )
            datos["Premio"] = premio if premio else "--"

            # ---------------- Suma Asegurada ----------------
            suma_asegurada = (
                buscar(r"Suma Asegurada\s*:\s*\$?\s*([0-9.]+,[0-9]{2})") or
                buscar(r"Suma Asegurada\s*:\s*\$?\s*([0-9.]+)")
            )
            datos["Suma Asegurada"] = suma_asegurada if suma_asegurada else "--"

            # ---------------- Cobertura (plan) ----------------
            plan = _extraer_plan_mercantil(texto_completo)
            if plan:
                datos["Cobertura"] = plan

        filas.append({col: datos[col] for col in columnas})

    nombre_archivo = excel_output_path("mercantil.xlsx")
    if os.path.exists(nombre_archivo):
        df_existente = pd.read_excel(nombre_archivo)
        df = pd.concat([df_existente, pd.DataFrame(filas)], ignore_index=True)
    else:
        df = pd.DataFrame(filas)

    df.to_excel(nombre_archivo, index=False)

    wb = load_workbook(nombre_archivo)
    ws = wb.active
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for cell in ws[1]:
        cell.fill = fill

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = 60 if col[0].value == "Cobertura" else max_len + 2

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
        ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

    wb.save(nombre_archivo)
    print(f"✅ Excel generado o actualizado como {nombre_archivo}")
