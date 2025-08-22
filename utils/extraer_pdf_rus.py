import pdfplumber
import pandas as pd
import re
import os
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from utils.resources import asset_path, excel_output_path

# --- Config detección de cláusula ---
ALLOWED_PCTS = {"10.0%", "15.0%", "20.0%", "25.0%", "30.0%", "40.0%"}
PCT_RE = r"\b(10(?:\.0)?|15(?:\.0)?|20(?:\.0)?|25(?:\.0)?|30(?:\.0)?|40(?:\.0)?)\s*%"

def _normalize_pct(raw: str) -> str:
    """Normaliza 10% -> 10.0% para comparar con ALLOWED_PCTS."""
    val = raw.strip()
    if not val.endswith("%"):
        val += "%"
    if "." not in val:
        val = val.replace("%", ".0%")
    return val

def _pick_allowed_pct_in_text(texto: str) -> str:
    """Busca un porcentaje permitido dentro de un texto ya acotado (p.ej., Cobertura)."""
    m = re.search(PCT_RE, texto, re.IGNORECASE)
    if not m:
        return ""
    val = _normalize_pct(m.group(1) + "%")
    return val if val in ALLOWED_PCTS else ""

def _pick_keyword_pct(texto: str) -> str:
    """
    Busca cerca de keywords típicas (incrementarán, automáticamente, ajuste, cláusula).
    No hace búsqueda global sin keywords para evitar falsos positivos como 83%.
    """
    t = " ".join(texto.split())
    m = re.search(
        rf"(?:incrementar(?:án)?|autom[aá]ticamente|ajuste|cl[aá]usula)[^%]{{0,120}}(?:hasta\s+un\s+)?{PCT_RE}",
        t, re.IGNORECASE
    )
    if m:
        val = _normalize_pct(m.group(1) + "%")
        if val in ALLOWED_PCTS:
            return val
    return ""   # <- sin fallback global

# --------- NUEVO: helpers para recortar cobertura ----------
def _compact_line(s: str) -> str:
    """Colapsa espacios/altos de línea para comparar y mostrar prolijo."""
    return re.sub(r"\s+", " ", s.strip())

def _cobertura_corta(cobertura_texto: str) -> str:
    """
    Devuelve solo las líneas relevantes del bloque:
      - Encabezados de cobertura (RCA/RCM/RCE/RCT ...)
      - Líneas de plan B-XX ...
      - 'S - Sigma'
    Mantiene orden y evita duplicados.
    """
    if not cobertura_texto:
        return ""

    lines = [_compact_line(l) for l in cobertura_texto.splitlines() if l.strip()]
    keep = []
    seen = set()

    for l in lines:
        # Encabezados de cobertura (tomar la línea completa)
        if re.match(r'^(RCA|RCM|RCE|RCT)\b', l, re.IGNORECASE):
            if l not in seen:
                keep.append(l); seen.add(l)
            continue

        # Planes B-XX ... (entera, con STD si aparece)
        if re.match(r'^B-\s*\d+\b.*', l, re.IGNORECASE):
            if l not in seen:
                keep.append(l); seen.add(l)
            continue

        # S - Sigma
        if re.match(r'^[Ss]\s*-\s*Sigma$', l, re.IGNORECASE):
            norm = "S - Sigma"
            if norm not in seen:
                keep.append(norm); seen.add(norm)
            continue

    # Si no se detectó nada, volver al texto original
    return "\n".join(keep) if keep else cobertura_texto
# ----------------------------------------------------------

def procesar_rus(pdfs: list[str]):
    columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura", "Archivo"]
    filas = []

    for pdf_path in pdfs:
        datos = {
            "Marca": "",
            "Modelo": "",
            "Año": "",
            "Suma Asegurada": "",
            "Premio": "--",
            "Cláusula de Ajuste": "--",
            "Cobertura": "--",
            "Archivo": os.path.basename(pdf_path)
        }

        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                extracted = page.extract_text() or ""
                texto_completo += extracted + "\n"

            def buscar(patron, multilinea=True):
                flags = re.MULTILINE | re.IGNORECASE if multilinea else re.IGNORECASE
                resultado = re.search(patron, texto_completo, flags)
                return resultado.group(1).strip() if resultado else ""

            # Marca, Modelo, Año
            match_vehiculo = re.search(
                r"Marca y modelo[:.\s]+([A-ZÁÉÍÓÚÑ]+)\s+(.+?)\s+Año[:.\s]+(\d{4})",
                texto_completo,
                re.IGNORECASE
            )
            if match_vehiculo:
                datos["Marca"] = match_vehiculo.group(1).strip()
                datos["Modelo"] = match_vehiculo.group(2).strip()
                datos["Año"] = match_vehiculo.group(3).strip()

            # Suma Asegurada
            datos["Suma Asegurada"] = buscar(r"Valor de reposición hasta la suma de[:.\s]+\$?\s*([0-9.\,]+)")

            # Premio
            datos["Premio"] = buscar(r"Premio\s*[:.]*\s*\$?\s*([0-9.\,]+)")

            # Cobertura: acotar el bloque de "Riesgos Cubiertos"
            cobertura_match = re.search(
                r"Riesgos Cubiertos\s*:?([\s\S]+?)(?=El Asegurador indemnizará|AUXILIO MECÁNICO|Advertencia al Asegurado|CARACTERÍSTICAS Y CONDICIONES|Cláusulas|CUIT|Frente de Póliza|$)",
                texto_completo,
                re.IGNORECASE
            )
            if cobertura_match:
                cobertura_texto = cobertura_match.group(1).strip()
                cobertura_texto = re.sub(r"\n{2,}", "\n", cobertura_texto)
                cobertura_texto = re.sub(r"[ \t]{2,}", " ", cobertura_texto)

                # === NUEVO: quedarnos solo con las líneas clave
                datos["Cobertura"] = _cobertura_corta(cobertura_texto)

            # Cláusula de ajuste: 1) dentro de Cobertura 2) cerca de keywords
            clau = ""
            if datos["Cobertura"]:
                clau = _pick_allowed_pct_in_text(datos["Cobertura"])
            if not clau:
                clau = _pick_keyword_pct(texto_completo)
            datos["Cláusula de Ajuste"] = clau if clau else "--"

        filas.append({col: datos[col] for col in columnas})

    nombre_archivo = excel_output_path("rio_uruguay.xlsx")

    if os.path.exists(nombre_archivo):
        df_existente = pd.read_excel(nombre_archivo)
        df = pd.concat([df_existente, pd.DataFrame(filas)], ignore_index=True)
    else:
        df = pd.DataFrame(filas)

    df.to_excel(nombre_archivo, index=False)

    wb = load_workbook(nombre_archivo)
    ws = wb.active

    # Fondo verde claro en encabezado
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    for cell in ws[1]:
        cell.fill = fill

    # Ajuste de columnas y estilo centrado
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = 60 if col[0].value == "Cobertura" else max_len + 2

    # Ajuste de altura de filas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
        ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

    wb.save(nombre_archivo)
    print(f"✅ Excel generado o actualizado como {nombre_archivo}")
