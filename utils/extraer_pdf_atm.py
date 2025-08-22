# utils/extraer_pdf_atm.py
import pdfplumber
import pandas as pd
import re
import os
import json
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
from utils.resources import asset_path, excel_output_path

with open(asset_path("assets", "marcas.json"), "r", encoding="utf-8") as f:
    marcas_json = json.load(f)
marcas_lista = [m["marca"].upper() for m in marcas_json]

def buscar(texto, patron, multilinea=True):
    flags = re.MULTILINE | re.IGNORECASE if multilinea else re.IGNORECASE
    resultado = re.search(patron, texto, flags)
    return resultado.group(1).strip() if resultado else ""

# --- Helper para convertir "231.766,69" -> 231766.69 (float) (por si lo necesitás) ---
def _to_float_ar(num_str: str) -> float:
    return float(num_str.replace('.', '').replace(',', '.'))

# --- FIX: detectar Premio por línea (misma línea que 'PREMIO' o las 2 siguientes) ---
_MONEY_RE = r"([0-9]{1,3}(?:\.[0-9]{3})*,[0-9]{2})"

def _premio_atm(texto: str) -> str:
    """
    Extrae el importe del PREMIO evitando confundir con SUBTOTAL/otros importes.
    1) Busca el label PREMIO por líneas y toma el último importe de esa línea.
    2) Si no hay en la misma línea, mira las 1-2 líneas siguientes.
    3) Fallback: busca a la derecha del label en una ventana corta.
    """
    lines = texto.splitlines()
    for i, line in enumerate(lines):
        if re.search(r"\bPREMIO(?:\s*DEL\s*PER[IÍ]ODO)?\b", line, re.IGNORECASE):
            # 1) Misma línea
            m_same = re.findall(_MONEY_RE, line)
            if m_same:
                return m_same[-1]  # el de la derecha si hay varios

            # 2) Siguientes 1-2 líneas (por si el importe quedó en la línea de abajo)
            for j in (i+1, i+2):
                if j < len(lines):
                    cand = re.findall(_MONEY_RE, lines[j])
                    if cand:
                        return cand[-1]

            # 3) Fallback corto a la derecha del label
            #    (evita agarrar montos grandes de arriba/otras secciones)
            full = "\n".join(lines)
            pos = full.find(line)
            if pos != -1:
                # posición del label dentro del texto completo
                label_match = re.search(r"\bPREMIO(?:\s*DEL\s*PER[IÍ]ODO)?\b", full[pos:pos+len(line)], re.IGNORECASE)
                if label_match:
                    end = pos + label_match.end()
                    right = full[end:end+180]
                    cand2 = re.findall(_MONEY_RE, right)
                    if cand2:
                        return cand2[0]
    return "--"

def procesar_atm(pdfs):
    filas = []

    for pdf_path in pdfs:
        datos = {
            "Marca": "",
            "Modelo": "",
            "Año": "",
            "Suma Asegurada": "--",
            "Premio": "--",
            "Cláusula de Ajuste": "--",
            "Cobertura": "--",
            "Archivo": os.path.basename(pdf_path)
        }

        with pdfplumber.open(pdf_path) as pdf:
            texto_completo = "\n".join([p.extract_text() for p in pdf.pages if p.extract_text()])

            # Marca y modelo
            marca_modelo = buscar(texto_completo, r"MARCA/MODELO:\s+([^\n]+)")
            if marca_modelo:
                texto = marca_modelo.upper()
                for marca_posible in sorted(marcas_lista, key=len, reverse=True):
                    if texto.startswith(marca_posible):
                        datos["Marca"] = marca_posible.title()
                        datos["Modelo"] = texto[len(marca_posible):].strip().title()
                        break

            # Año
            datos["Año"] = buscar(texto_completo, r"AÑO:\s*(\d{4})")

            # Suma asegurada
            suma = buscar(texto_completo, r"SUMA ASEGURADA:\s*([0-9.]+,[0-9]{2})") or \
                   buscar(texto_completo, r"SUMA ASEGURADA:\s*([0-9.]+)")
            datos["Suma Asegurada"] = suma if suma else "--"

            # Cláusula de ajuste
            ajuste = buscar(texto_completo, r"CLAUSULA DE AJUSTE AUTOMATICO\s*:\s*(\d+ ?%)")
            datos["Cláusula de Ajuste"] = ajuste if ajuste else "--"

            # Cobertura
            cobertura = buscar(texto_completo, r"COBERTURA:\s*([^\n]+)")
            datos["Cobertura"] = cobertura if cobertura else "--"

            # --- PREMIO corregido (línea del label / líneas siguientes / fallback corto) ---
            datos["Premio"] = _premio_atm(texto_completo)

        filas.append(datos)

    columnas = ["Marca", "Modelo", "Año", "Suma Asegurada", "Premio", "Cláusula de Ajuste", "Cobertura", "Archivo"]
    df = pd.DataFrame(filas, columns=columnas)
    nombre_archivo = excel_output_path("atm.xlsx")
    df.to_excel(nombre_archivo, index=False)

    wb = load_workbook(nombre_archivo)
    ws = wb.active
    fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    for cell in ws[1]:
        cell.fill = fill

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[col_letter].width = 60 if col[0].value == "Cobertura" else max_len + 2

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        max_lines = max(str(c.value).count("\n") + 1 if c.value else 1 for c in row)
        ws.row_dimensions[row[0].row].height = max(15, max_lines * 15)

    wb.save(nombre_archivo)
    print(f"✅ Excel generado o actualizado como {nombre_archivo}")
